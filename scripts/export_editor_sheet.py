#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CONTENT_PATH = ROOT / "content" / "prante-content.json"
SCHEMA_PATH = ROOT / "content" / "editor-schema.json"
CSV_OUTPUT_PATH = ROOT / "content" / "editor-texts-template.csv"
EXCEL_CSV_OUTPUT_PATH = ROOT / "content" / "editor-texts-template.excel.csv"
XLSX_OUTPUT_PATH = ROOT / "content" / "editor-texts-template.xlsx"
HEADERS = ["Раздел", "Поле", "Текст", "Подсказка", "Лимит"]


def parse_path(path: str) -> list[str | int]:
    parts: list[str | int] = []
    token = ""
    index = ""
    in_index = False
    for char in path:
        if char == "." and not in_index:
            if token:
                parts.append(token)
                token = ""
        elif char == "[":
            if token:
                parts.append(token)
                token = ""
            in_index = True
            index = ""
        elif char == "]":
            parts.append(int(index))
            in_index = False
        elif in_index:
            index += char
        else:
            token += char
    if token:
        parts.append(token)
    return parts


def get_path(data: object, path: str) -> object:
    current = data
    for part in parse_path(path):
        current = current[part]  # type: ignore[index]
    return current


def build_rows() -> list[dict[str, object]]:
    content = json.loads(CONTENT_PATH.read_text(encoding="utf-8"))
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))["fields"]

    rows: list[dict[str, object]] = []
    for field in schema:
        rows.append(
            {
                "Раздел": field.get("section", ""),
                "Поле": field.get("label", ""),
                "Текст": get_path(content, field["path"]),
                "Подсказка": field.get("notes", ""),
                "Лимит": field.get("maxLength", ""),
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, object]], delimiter: str) -> bool:
    try:
        with path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=HEADERS, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(rows)
    except PermissionError:
        print(f"Skipped locked file {path}", file=sys.stderr)
        return False
    return True


def write_xlsx(path: Path, rows: list[dict[str, object]]) -> bool:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Тексты"

    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row[header] for header in HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    text_fill = PatternFill("solid", fgColor="E2F0D9")
    helper_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(color="FFFFFF", bold=True)
    top_wrap = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 22,
        "B": 30,
        "C": 78,
        "D": 54,
        "E": 10,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = top_wrap
        row[2].fill = text_fill
        row[3].fill = helper_fill
        row[4].fill = helper_fill

    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 42

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    sheet["C1"].comment = Comment("Редакторы меняют только эту колонку.", "PRANTE")
    sheet["D1"].comment = Comment("Подсказка объясняет контекст, её не нужно переносить в интерфейс.", "PRANTE")

    guide = workbook.create_sheet("Инструкция")
    guide_rows = [
        ["Как редактировать тексты ПРАНТЕ"],
        ["1. Меняйте только колонку «Текст» на листе «Тексты»."],
        ["2. Колонки «Раздел», «Поле», «Подсказка» и «Лимит» помогают понять контекст."],
        ["3. Не начинайте стартовое сообщение с «Здравствуйте»: приветствие добавляется отдельно."],
        ["4. После правок проверьте preview в /admin/."],
        ["5. Если строка стала слишком длинной, preview покажет предупреждение."],
    ]
    for guide_row in guide_rows:
        guide.append(guide_row)
    guide.column_dimensions["A"].width = 96
    guide["A1"].font = Font(bold=True, size=14, color="1F4E79")
    for row in guide.iter_rows():
        row[0].alignment = Alignment(vertical="top", wrap_text=True)

    try:
        workbook.save(path)
    except PermissionError:
        print(f"Skipped locked file {path}", file=sys.stderr)
        return False
    return True


def main() -> None:
    rows = build_rows()

    if write_csv(CSV_OUTPUT_PATH, rows, ","):
        print(f"Wrote {CSV_OUTPUT_PATH}")
    if write_csv(EXCEL_CSV_OUTPUT_PATH, rows, ";"):
        print(f"Wrote {EXCEL_CSV_OUTPUT_PATH}")
    if write_xlsx(XLSX_OUTPUT_PATH, rows):
        print(f"Wrote {XLSX_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
