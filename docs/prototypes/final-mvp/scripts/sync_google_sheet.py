#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import urllib.request
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CONTENT_PATH = ROOT / "content" / "prante-content.json"
SCHEMA_PATH = ROOT / "content" / "editor-schema.json"
CONFIG_PATH = ROOT / "content" / "editor-config.json"


def parse_path(path: str) -> list[str | int]:
    parts: list[str | int] = []
    token = ""
    index = ""
    in_index = False
    for char in path:
        if char == "." and not in_index:
            if token:
                parts.append(token)
                token = ""
        elif char == "[":
            if token:
                parts.append(token)
                token = ""
            in_index = True
            index = ""
        elif char == "]":
            parts.append(int(index))
            in_index = False
        elif in_index:
            index += char
        else:
            token += char
    if token:
        parts.append(token)
    return parts


def get_path(data: object, path: str) -> object:
    current = data
    for part in parse_path(path):
        current = current[part]  # type: ignore[index]
    return current


def set_path(data: object, path: str, value: object) -> None:
    current = data
    parts = parse_path(path)
    for index, part in enumerate(parts):
        last = index == len(parts) - 1
        if last:
            current[part] = value  # type: ignore[index]
            return
        current = current[part]  # type: ignore[index]


def parse_csv_rows(text: str) -> list[dict[str, str]]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    return list(csv.DictReader(text.splitlines(), dialect=dialect))


def load_xlsx_rows(source: str) -> list[dict[str, str]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["Тексты"] if "Тексты" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    result: list[dict[str, str]] = []
    for row in rows:
        item = {
            headers[index]: "" if value is None else str(value)
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(value.strip() for value in item.values()):
            result.append(item)
    return result


def load_rows(source: str) -> list[dict[str, str]]:
    if source.startswith("http://") or source.startswith("https://"):
        with urllib.request.urlopen(source, timeout=20) as response:
            text = response.read().decode("utf-8-sig")
        return parse_csv_rows(text)

    path = Path(source)
    if path.suffix.lower() == ".xlsx":
        return load_xlsx_rows(source)
    else:
        text = path.read_text(encoding="utf-8-sig")
        return parse_csv_rows(text)


def row_key(row: dict[str, str]) -> tuple[str, str]:
    return (
        (row.get("section") or row.get("Section") or row.get("Раздел") or "").strip().lower(),
        (row.get("label") or row.get("Label") or row.get("Поле") or row.get("Название") or "").strip().lower(),
    )


def row_path(row: dict[str, str], by_label: dict[tuple[str, str], dict]) -> str:
    direct = (row.get("path") or row.get("Path") or row.get("Ключ") or row.get("Путь") or "").strip()
    if direct:
        return direct
    field = by_label.get(row_key(row))
    return field["path"] if field else ""


def row_value(row: dict[str, str]) -> str | None:
    if "value" in row:
        return row.get("value")
    if "Value" in row:
        return row.get("Value")
    return row.get("Текст")


def validate(content: dict, schema: list[dict]) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    for field in schema:
        value = str(get_path(content, field["path"]) or "").strip()
        if field.get("required") and not value:
            errors.append(f"{field['path']}: empty required value")
        max_length = field.get("maxLength")
        if max_length and len(value) > int(max_length):
            warnings.append(f"{field['path']}: {len(value)}/{max_length} chars")

    start = str(get_path(content, "chat.startMessage") or "").strip()
    if re.match(r"^здравствуйте[.!?,\s]", start, re.IGNORECASE):
        errors.append("chat.startMessage: duplicate greeting risk")

    return errors, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync PRANTE editorial sheet into fallback JSON.")
    parser.add_argument("--source", help="Published Google Sheet CSV URL, local CSV path, or local XLSX path. Defaults to editor-config.json.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and print summary without writing JSON.")
    args = parser.parse_args()

    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    source = args.source or config.get("sheetCsvUrl")
    if not source:
        print("No source provided. Pass --source or set content/editor-config.json sheetCsvUrl.", file=sys.stderr)
        return 2

    content = json.loads(CONTENT_PATH.read_text(encoding="utf-8"))
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))["fields"]
    allowed = {field["path"]: field for field in schema}
    by_label = {
        ((field.get("section") or "").strip().lower(), (field.get("label") or "").strip().lower()): field
        for field in schema
    }
    rows = load_rows(source)
    merged = deepcopy(content)
    applied = 0
    skipped = 0

    for row in rows:
        path = row_path(row, by_label)
        value = row_value(row)
        if not path or path not in allowed:
            skipped += 1
            continue
        if value is None or not str(value).strip():
            skipped += 1
            continue
        set_path(merged, path, str(value))
        applied += 1

    errors, warnings = validate(merged, schema)
    for warning in warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        return 1

    print(f"Rows applied: {applied}; skipped: {skipped}; warnings: {len(warnings)}")
    if not args.dry_run:
        CONTENT_PATH.write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"Updated {CONTENT_PATH}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
